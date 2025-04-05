#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings

def demonstrate_excel_handling():
    """
    Demonstrate different approaches to handling Excel files with various complexities.
    This example shows common scenarios and their solutions.
    """


# In[ ]:


# Part 1: Basic Excel Reading
print("Part 1: Basic Excel Reading")
print("-" * 50)

try:
    # Basic reading of an Excel file
    df_basic = pd.read_excel('sales_data.xlsx')
    print("Basic Excel read successful")
    print(df_basic.head())
except FileNotFoundError:
    print("Creating example Excel file...")
    
    # Create sample data
    data = {
        'Date': pd.date_range('2024-01-01', periods=10),
        'Product': ['A', 'B', 'C', 'A', 'B', 'C', 'A', 'B', 'C', 'A'],
        'Sales': np.random.randint(100, 1000, 10),
        'Region': ['North', 'South', 'East', 'West', 'North', 'South', 'East', 'West', 'North', 'South']
    }
    df_sample = pd.DataFrame(data)
    df_sample.to_excel('sales_data.xlsx', sheet_name='Sales', index=False)
    print("Sample Excel file created")


# In[ ]:


# Part 2: Working with Multiple Sheets
print("\nPart 2: Working with Multiple Sheets")
print("-" * 50)

# Create a multi-sheet Excel file
with pd.ExcelWriter('multi_sheet_sales.xlsx') as writer:
    # Create different sheets for each region
    for region in ['North', 'South', 'East', 'West']:
        df_region = df_sample[df_sample['Region'] == region]
        df_region.to_excel(writer, sheet_name=region, index=False)
        
    # Create a summary sheet with different formatting
    summary_data = df_sample.groupby('Region')['Sales'].agg(['sum', 'mean', 'count'])
    summary_data.to_excel(writer, sheet_name='Summary')

# Reading specific sheets
def read_multiple_sheets():
    # Read all sheets into a dictionary
    all_sheets = pd.read_excel('multi_sheet_sales.xlsx', sheet_name=None)
    print("Available sheets:", list(all_sheets.keys()))
    
    # Read specific sheets
    north_data = pd.read_excel('multi_sheet_sales.xlsx', sheet_name='North')
    print("\nNorth region data:")
    print(north_data.head())
    
    return all_sheets

sheets_dict = read_multiple_sheets()


# In[ ]:


# Part 3: Handling Merged Cells and Formatted Data
print("\nPart 3: Handling Merged Cells and Formatted Data")
print("-" * 50)

def create_formatted_excel():
    """Create an Excel file with merged cells and formatting"""
    # Create a workbook with formatted data
    wb = load_workbook('multi_sheet_sales.xlsx')
    sheet = wb.create_sheet('Formatted')
    
    # Add merged cells for headers
    sheet.merge_cells('A1:D1')
    sheet['A1'] = 'Sales Report 2024'
    
    # Add subheaders
    headers = ['Date', 'Product', 'Sales', 'Region']
    for col, header in enumerate(headers, 1):
        sheet.cell(row=2, column=col, value=header)
    
    # Add some sample data
    sample_data = sheets_dict['North'].values.tolist()
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save('formatted_sales.xlsx')
    print("Created formatted Excel file")

create_formatted_excel()


# In[ ]:


# Reading Excel with merged cells
def read_formatted_excel():
    """Demonstrate reading Excel with merged cells"""
    # Using openpyxl engine
    try:
        df_formatted = pd.read_excel('formatted_sales.xlsx', 
                                   sheet_name='Formatted',
                                   header=1,  # Skip merged header
                                   engine='openpyxl')
        print("\nRead formatted data with openpyxl engine:")
        print(df_formatted.head())
    except Exception as e:
        print(f"Error reading with openpyxl: {e}")
    
    # Using xlrd engine for comparison (if .xls file)
    try:
        df_formatted_xlrd = pd.read_excel('formatted_sales.xlsx',
                                        sheet_name='Formatted',
                                        header=1,
                                        engine='xlrd')
        print("\nRead formatted data with xlrd engine:")
        print(df_formatted_xlrd.head())
    except Exception as e:
        print(f"Note: xlrd engine only supports .xls files: {e}")
        
read_formatted_excel()


# In[ ]:


# Part 4: Advanced Excel Handling
print("\nPart 4: Advanced Excel Handling")
print("-" * 50)

def demonstrate_advanced_features():
    """Show advanced Excel handling features"""
    # Reading specific columns
    df_cols = pd.read_excel('sales_data.xlsx',
                          usecols=['Date', 'Sales'],
                          converters={'Sales': lambda x: float(x) if pd.notnull(x) else 0})
    print("\nReading specific columns with converters:")
    print(df_cols.head())
    
    # Reading with custom NA values
    df_na = pd.read_excel('sales_data.xlsx',
                        na_values=['NA', 'missing', -999],
                        keep_default_na=False)
    print("\nReading with custom NA values:")
    print(df_na.head())
    
    # Reading with date parsing
    df_dates = pd.read_excel('sales_data.xlsx',
                           parse_dates=['Date'],
                           date_parser=lambda x: pd.to_datetime(x, format='%Y-%m-%d'))
    print("\nReading with date parsing:")
    print(df_dates.head())
    
demonstrate_advanced_features()

